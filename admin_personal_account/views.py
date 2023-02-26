from django.core.paginator import Paginator
from django.db.models import Value
from django.db.models.functions import Concat
from django.http import JsonResponse, HttpResponse
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.views.generic import CreateView, UpdateView, ListView, DetailView
from django.contrib import messages
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

from admin_house.models import House
from admin_transaction.models import Transaction
from users.mixins import AdminPermissionRequiredMixin
from .models import PersonalAccount
from .forms import PersonalAccountForm
from admin_apartment.models import Apartment, Section
from openpyxl import Workbook

from .utils import PersonalAccountBalance


# Create your views here.

class PersonalAccountView(AdminPermissionRequiredMixin, DetailView):
    permission_required = 'personal_accounts'
    model = PersonalAccount
    template_name = 'admin_personal_account/personal_account_view.html'


class PersonalAccountCreate(AdminPermissionRequiredMixin, CreateView):
    permission_required = 'personal_accounts'
    model = PersonalAccount
    form_class = PersonalAccountForm
    template_name = 'admin_personal_account/personal_account_create.html'
    success_url = reverse_lazy('personal_account_list')

    def render_to_response(self, context, **response_kwargs):
        if self.request.is_ajax():
            house_id = self.request.GET.get('house_id')
            section_id = self.request.GET.get('section_id')
            print(self.request.GET)
            if house_id:
                if section_id:
                    result = {'apartment': list(Apartment.objects.filter(
                        house=house_id, section=section_id).values('id', 'number', 'owner__first_name', 'owner',
                                                                   'owner__middle_name', 'owner__last_name',
                                                                   'owner__phone'))}
                else:
                    result = {'apartment': list(Apartment.objects.filter(
                        house_id=house_id, personalaccount__isnull=True).values('id', 'number', 'owner__first_name',
                                                                                'owner__phone', 'owner__middle_name',
                                                                                'owner__last_name', 'owner')),
                              'section': list(Section.objects.filter(house_id=house_id).values())}
                print(result)
                return JsonResponse(result, safe=False, **response_kwargs)
        else:
            return super(CreateView, self).render_to_response(context, **response_kwargs)


class PersonalAccountUpdate(AdminPermissionRequiredMixin, UpdateView):
    permission_required = 'personal_accounts'
    model = PersonalAccount
    template_name = 'admin_personal_account/personal_account_update.html'
    form_class = PersonalAccountForm
    success_url = reverse_lazy('personal_account_list')


def personal_account_delete(request, pk):
    number = None
    try:
        obj_delete = PersonalAccount.objects.get(pk=pk)
        n = obj_delete.__str__()
        if obj_delete.delete():
            number = n
    except:
        messages.error(request, f"Не удалось удалить лицевой счет")
    if number:
        messages.success(request, f"Лицевой счет №{number} удален успешно")
    return redirect('personal_account_list')


class PersonalAccountList(AdminPermissionRequiredMixin, ListView):
    permission_required = 'personal_accounts'
    model = PersonalAccount
    template_name = 'admin_personal_account/personal_account_list.html'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(PersonalAccountList, self).get_context_data(**kwargs)
        context['house_list'] = House.objects.all()
        context['account_total_debt'] = PersonalAccountBalance.get_total_debt()
        context['account_total_balance'] = PersonalAccountBalance.get_total_balance()
        context['transactions_total_balance'] = Transaction.total_balance()
        return context

    def to_excel(self, value_list):
        wb = Workbook()
        ws = wb.active
        title_list = ['Лицевой счет', 'Статус', 'Дом', 'Секция', 'Квартира', 'Владелец', 'Остаток']
        ws.append(title_list)
        for account in value_list:
            ws.append(list(account.values())[1:])
        for i in range(1, len(title_list) + 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 20
        response = HttpResponse(save_virtual_workbook(wb),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=export.xlsx'
        return response

    def render_to_response(self, context, **response_kwargs):
        if self.request.is_ajax():
            print(self.request.GET)
            result = {}
            filter_fields = {
                'number__contains': self.request.GET.get('number'),
                'status': self.request.GET.get('status'),
                'apartment__house': self.request.GET.get('house'),
                'apartment__section_id': self.request.GET.get('section'),
                'apartment__number__contains': self.request.GET.get('apartment'),
                'apartment__owner': self.request.GET.get('owner'),
            }
            if self.request.GET.get('house_select'):
                result['sections'] = list(Section.objects.filter(house=filter_fields['apartment__house']).values())

            filter_fields = {k: v for k, v in filter_fields.items() if v}
            PersonalAccount.objects.select_related('')
            filtered_qs = self.get_queryset().prefetch_related('receipt_set', 'transaction_set').filter(**filter_fields)
            filtered_qs = filtered_qs.annotate(
                owner__name=Concat('apartment__owner__first_name', Value(' '),
                                   'apartment__owner__middle_name', Value(' '),
                                   'apartment__owner__last_name'))
            data = filtered_qs.values('id', 'number', 'status', 'apartment__house__name',
                                      'apartment__section__name', 'apartment__number', 'owner__name')
            start = int(self.request.GET.get('start', 0))
            length = int(self.request.GET.get('length', 10))
            paginator = Paginator(data, self.request.GET.get('length', 10))
            page = (start // length) + 1
            data = list(paginator.get_page(page))
            for personal_account in data:
                personal_account['balance'] = filtered_qs.get(id=personal_account['id']).balance

            result['account'] = data
            result['recordsTotal'] = paginator.count
            result['recordsFiltered'] = paginator.count
            result['pages'] = paginator.num_pages
            print(result)
            if self.request.GET.get('to_excel'):
                return self.to_excel(value_list=result['account'])
            return JsonResponse(result, safe=False, **response_kwargs)
        else:
            return super(ListView, self).render_to_response(context, **response_kwargs)
