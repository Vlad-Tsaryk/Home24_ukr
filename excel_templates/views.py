from copy import copy

from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from django.core.mail import EmailMessage
from django.http import HttpResponse
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import CreateView, DeleteView, TemplateView
from django.views.generic.detail import SingleObjectMixin
from openpyxl.cell import Cell
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.writer.excel import save_virtual_workbook

from admin_purpose.models import PaymentDetails
from admin_receipt.models import Receipt
from excel_templates.models import ExcelTemplate
from home24.settings import EMAIL_HOST
from users.mixins import AdminPermissionRequiredMixin
from .forms import ExcelTemplateCreateForm


# Create your views here.
class ExcelTemplateCreate(AdminPermissionRequiredMixin, SuccessMessageMixin, CreateView):
    permission_required = 'receipts'
    model = ExcelTemplate
    form_class = ExcelTemplateCreateForm
    template_name = 'excel_templates/excel_template_create.html'
    success_message = 'Шаблон %(name)s успішно добавлен'
    success_url = reverse_lazy('excel-template-create')

    def get_context_data(self, **kwargs):
        context = super(ExcelTemplateCreate, self).get_context_data(**kwargs)
        context['excel_template_list'] = ExcelTemplate.objects.all().order_by('-pk')
        return context


class ExcelTemplateDelete(AdminPermissionRequiredMixin, DeleteView):
    permission_required = 'receipts'
    model = ExcelTemplate

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        messages.success(self.request, f'Шаблон {obj.name} успішно видалено')
        if obj.default and ExcelTemplate.objects.count() > 1:

            self.success_url = reverse_lazy('excel-template-set-default',
                                            kwargs={'pk': ExcelTemplate.objects.exclude(pk=obj.pk).last().pk})
        else:
            self.success_url = reverse_lazy('excel-template-create')
        return super(ExcelTemplateDelete, self).delete(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        return self.delete(request, *args, **kwargs)


class ExcelTemplateSetDefault(AdminPermissionRequiredMixin, SingleObjectMixin, View):
    permission_required = 'receipts'
    model = ExcelTemplate

    def get(self, request, *args, **kwargs):
        obj = self.get_object()
        try:
            previous_default_obj = ExcelTemplate.objects.get(default=True)
            previous_default_obj.default = False
            previous_default_obj.save()
        except:
            pass
        obj.default = True
        obj.save()
        messages.info(self.request, f'Шаблон {obj.name} задан по умолчанию')
        return redirect('excel-template-create')


class ExcelTemplatePrint(AdminPermissionRequiredMixin, SingleObjectMixin, TemplateView):
    permission_required = 'receipts'
    model = Receipt
    template_name = 'excel_templates/excel_template_print.html'

    def get_context_data(self, **kwargs):
        context = super(SingleObjectMixin, self).get_context_data(**kwargs)
        context['receipt'] = self.get_object()
        context['excel_template_list'] = ExcelTemplate.objects.all().order_by('-pk')
        print(context)
        return context

    def post(self, request, *args, **kwargs):
        obj = self.get_object()
        excel_template = ExcelTemplate.objects.get(pk=self.request.POST.get('template_id'))
        print(excel_template)
        print(obj)
        if self.request.POST.get('action_send_email'):
            if ReceiptToExcel(obj, 'action_send_email', excel_template).get_excel():
                messages.success(request, 'Email отправлен успішно')
            return self.render_to_response(self.get_context_data())
        elif self.request.POST.get('action_download'):
            return ReceiptToExcel(obj, 'action_download', excel_template).get_excel()


class ReceiptToExcel:
    def __init__(self, receipt_object, response_type, excel_template, **kwargs):
        self.receipt_object = receipt_object
        self.response_type = response_type
        self.excel_template = excel_template.file
        for key, value in kwargs.items():
            setattr(self, key, value)

    def get_excel(self):
        ru_months = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                     'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
        account_balance = self.receipt_object.personal_account.balance
        receipt_selectors = {
            '$payCompany$': PaymentDetails.objects.first().name,
            '$receiptAddress$': self.receipt_object.address_for_excel,
            '$receiptNumber$': self.receipt_object.number,
            '$accountNumber$': self.receipt_object.personal_account.number,
            '$receiptDate$': self.receipt_object.date.strftime('%d.%m.%Y'),
            '$receiptMonth$': f'{ru_months[self.receipt_object.date.month - 1]} {self.receipt_object.date.year}',
            '$accountBalance$': account_balance,
            '$receiptPayable$': account_balance - self.receipt_object.total_price,
            '$total$': self.receipt_object.total_price,
        }
        receipt_services = self.receipt_object.receiptservice_set.all()

        def get_receipt_service_selectors(selector, receipt_service):
            receipt_service_selectors = {
                '$serviceName$': receipt_service.service.name,
                '$servicePrice$': receipt_service.price_unit,
                '$serviceAmount$': receipt_service.consumption,
                '$serviceUnit$': receipt_service.service.unit.name,
                '$serviceTotal$': receipt_service.price_unit * receipt_service.consumption,
            }
            return receipt_service_selectors.get(selector) or selector

        wb = load_workbook(self.excel_template)
        ws = wb.active
        loop_coord = None
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val in receipt_selectors:
                    cell.value = receipt_selectors[val]
                elif val == '$LOOP 1$':
                    loop_coord = cell.row
        if loop_coord:
            target_row = ws[loop_coord + 1]
            row_number = receipt_services.count()
            row_height = ws.row_dimensions[loop_coord].height
            ws.delete_rows(loop_coord)
            cell_range = CellRange(f"{target_row[0].coordinate}:{target_row[-1].coordinate}")
            for row in range(1, row_number + 1):
                for merged_cell in ws.merged_cells:
                    if merged_cell.coord not in cell_range:
                        continue
                    cr = CellRange(merged_cell.coord)
                    cr.shift(row_shift=row)
                    ws.merge_cells(cr.coord)
            if row_number > 1:
                ws.insert_rows(loop_coord + 1, amount=row_number - 1)
                for row in range(1, row_number):
                    ws.row_dimensions[loop_coord + row + 1].height = row_height
                    for cell in target_row:
                        new_cell = ws.cell(row=cell.row + row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell._style = copy(cell._style)

            for row, receipt_service in zip(
                    ws.iter_rows(min_row=loop_coord, max_row=loop_coord + row_number, max_col=target_row[-1].column),
                    receipt_services):
                for cell in row:
                    if isinstance(cell, Cell):
                        cell.value = get_receipt_service_selectors(cell.value, receipt_service)
        ws_save = save_virtual_workbook(wb)

        if self.response_type == 'action_send_email':
            email = EmailMessage()
            email.attach(f'receipt__{receipt_selectors["$receiptNumber$"]}_{receipt_selectors["$receiptDate$"]}.xlsx',
                         ws_save, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            email.from_email = EMAIL_HOST
            email.subject = 'Test'
            email.body = 'Test'
            email.to = [self.receipt_object.apartment.owner.username]
            if email.send():
                return True
            return False
        elif self.response_type == 'action_download':
            response = HttpResponse(ws_save, charset='utf-8',
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=receipt__{receipt_selectors["$receiptNumber$"]}_' \
                                              f'{receipt_selectors["$receiptDate$"]}.xlsx'

            return response

