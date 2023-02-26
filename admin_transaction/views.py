from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Value
from django.db.models.functions import Concat
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import CreateView, ListView, DetailView, UpdateView, DeleteView
from django.views.generic.detail import SingleObjectMixin
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from admin_personal_account.utils import PersonalAccountBalance
from admin_transaction.models import Transaction
from users.mixins import AdminPermissionRequiredMixin
from .forms import TransactionForm
from admin_personal_account.models import PersonalAccount
from users.models import User, Role
from admin_purpose.models import Purpose
from django.contrib.messages import success, error


# Create your views here.
class TransactionCreate(AdminPermissionRequiredMixin, CreateView):
    permission_required = 'transactions'
    model = Transaction
    form_class = TransactionForm
    template_name = 'admin_transaction/transaction_create.html'
    success_url = reverse_lazy('transaction_list')

    def get_form_kwargs(self):
        kwargs = super(TransactionCreate, self).get_form_kwargs()
        kwargs['transaction_type'] = self.kwargs.get('transaction_type')
        kwargs['user_id'] = self.request.user.pk
        return kwargs

    def render_to_response(self, context, **response_kwargs):
        if self.request.is_ajax():
            result = {}
            owner_id = self.request.GET.get('owner_id')
            personal_account_id = self.request.GET.get('personal_account')
            print(self.request.GET)
            if self.request.GET.get('owner_change'):
                if owner_id:
                    result['personal_accounts'] = list(PersonalAccount.objects.filter(
                        apartment__owner=owner_id).values('id', 'number'))
                else:
                    result['personal_accounts'] = list(PersonalAccount.objects.all().values('id', 'number'))

            if self.request.GET.get('personal_account_change'):
                if personal_account_id:
                    result['owner_id'] = str(PersonalAccount.objects.get(
                        pk=personal_account_id).apartment.owner_id)
            print(result)
            return JsonResponse(result, safe=False, **response_kwargs)
        else:
            return super(CreateView, self).render_to_response(context, **response_kwargs)


class TransactionUpdate(TransactionCreate, UpdateView):
    template_name = 'admin_transaction/transaction_update.html'
    model = Transaction
    form_class = TransactionForm
    success_url = reverse_lazy('transaction_list')


class TransactionClone(TransactionUpdate):
    template_name = 'admin_transaction/transaction_update.html'

    def get_form_kwargs(self):
        kwargs = super(TransactionClone, self).get_form_kwargs()
        transaction_obj = Transaction.objects.get(pk=self.kwargs['pk'])
        if transaction_obj:
            transaction_obj.pk = None
            transaction_obj.number = None
            kwargs['instance'] = transaction_obj
        return kwargs


class TransactionList(AdminPermissionRequiredMixin, ListView):
    permission_required = 'transactions'
    model = Transaction
    template_name = 'admin_transaction/transaction_list.html'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(TransactionList, self).get_context_data(**kwargs)
        context['owner_list'] = User.objects.filter(role__role=Role.RoleName.OWNER)
        context['account_total_debt'] = PersonalAccountBalance.get_total_debt()
        context['account_total_balance'] = PersonalAccountBalance.get_total_balance()
        context['transactions_total_balance'] = Transaction.total_balance()
        context['purpose_list'] = Purpose.objects.all()
        return context

    def to_excel(self, value_list):
        wb = Workbook()
        ws = wb.active
        title_list = ['№', 'Дата', 'Приход/расход', 'Статус', 'Статья', 'Сумма', 'Валюта', 'Лицевой счет',
                      'Владелец квартиры']
        ws.append(title_list)
        status_dict = {0: 'Не проведен', 1: 'Проведен'}
        type_dict = {0: 'Расход', 1: 'Приход'}
        for transaction in value_list:
            transaction_list = list(transaction.values())
            transaction_list[2] = type_dict[transaction_list[2]]
            transaction_list[3] = status_dict[transaction_list[3]]
            transaction_list[6] = 'UAH'
            ws.append(transaction_list)
        for i in range(1, len(title_list) + 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 20
        response = HttpResponse(save_virtual_workbook(wb),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=export.xlsx'
        return response

    def render_to_response(self, context, **response_kwargs):
        if self.request.is_ajax():
            print(self.request.GET)
            order_by = self.request.GET.get('order_by')
            result = {}
            filter_fields = {
                'status': self.request.GET.get('status'),
                'number__contains': self.request.GET.get('number'),
                'owner_id': self.request.GET.get('owner'),
                'type': self.request.GET.get('purpose_type'),
                'personal_account__number__contains': self.request.GET.get('personal_account__number'),
                'purpose': self.request.GET.get('purpose'),
                'is_complete': self.request.GET.get('is_complete'),
                'date__range': self.request.GET.getlist('date_range[]'),
            }
            # if self.request.GET.get('purpose_type_change'):
            #     if filter_fields['type']:
            #         result['purposes'] = list(
            #             context['purpose_list'].filter(transaction_type=filter_fields['type']).values('name'))
            #     else:
            #         result['purposes'] = list(context['purpose_list'].values('name'))
            filter_fields = {k: v for k, v in filter_fields.items() if v}
            filtered_qs = self.get_queryset().filter(**filter_fields)
            filtered_qs = filtered_qs.annotate(
                owner__name=Concat('owner__first_name', Value(' '),
                                   'owner__middle_name', Value(' '),
                                   'owner__last_name'))

            if order_by:
                filtered_qs = filtered_qs.order_by(order_by)
            result['income'] = '%.2f' % self.model.count_income(filtered_qs)
            result['outcome'] = '%.2f' % self.model.count_outcome(filtered_qs)
            filtered_qs = filtered_qs.values('number', 'date', 'type', 'is_complete', 'purpose__name',
                                             'sum', 'id', 'owner__name', 'personal_account__number', )
            start = int(self.request.GET.get('start', 0))
            length = int(self.request.GET.get('length', 10))
            paginator = Paginator(filtered_qs, self.request.GET.get('length', 10))
            page = (start // length) + 1
            data = list(paginator.get_page(page))
            result['transaction'] = data
            result['recordsTotal'] = paginator.count
            result['recordsFiltered'] = paginator.count
            result['pages'] = paginator.num_pages
            print(result)
            if self.request.GET.get('to_excel'):
                return self.to_excel(value_list=result['transaction'])
            return JsonResponse(result, safe=False, **response_kwargs)

        else:
            return super(ListView, self).render_to_response(context, **response_kwargs)


class TransactionView(AdminPermissionRequiredMixin, DetailView):
    permission_required = 'transactions'
    model = Transaction
    template_name = 'admin_transaction/transaction_view.html'


class TransactionDelete(AdminPermissionRequiredMixin, DeleteView):
    permission_required = 'transactions'
    model = Transaction
    success_url = reverse_lazy('transaction_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        success_url = self.get_success_url()
        obj_number = self.object.number
        try:
            if self.object.delete():
                messages.success(self.request, f"Ведомость №{obj_number} успешно удалена")
        except:
            messages.error(request, f"Не удалось удалить ведомость")
        return HttpResponseRedirect(success_url)

    def get(self, request, *args, **kwargs):
        return self.delete(request, *args, **kwargs)


class TransactionToExcel(View, SingleObjectMixin):
    model = Transaction

    def get(self, request, *args, **kwargs):
        transaction = self.get_object()
        wb = Workbook()
        ws = wb.active
        status_dict = {0: 'Не проведен', 1: 'Проведен'}
        type_dict = {0: 'Расход', 1: 'Приход'}
        date = transaction.date.strftime('%d.%m.%Y')
        number = transaction.number
        content = [['Платеж', number],
                   ['Дата', date],
                   ['Владелец квартиры', str(transaction.owner or '')],
                   ['Лицевой счет', str(transaction.personal_account or '')],
                   ['Приход/расход', type_dict[transaction.type]],
                   ['Статус', status_dict[transaction.is_complete]],
                   ['Статья', str(transaction.purpose)],
                   ['Сумма', transaction.sum],
                   ['Валюта', 'UAH'],
                   ['Комментарий', transaction.comment],
                   ['Менеджер', str(transaction.manager)]]
        for row_content in content:
            ws.append(row_content)
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        file_name = f'transaction__#{number}_{date}'
        response = HttpResponse(save_virtual_workbook(wb),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'
        return response
